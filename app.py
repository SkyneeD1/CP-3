import streamlit as st
import pandas as pd
from datetime import datetime, date, time, timedelta
from io import BytesIO
import re
from enum import Enum
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ----------------------------
# ENUMS e CONSTANTES
# ----------------------------
class TipoApuracao(Enum):
    DIARIA = "Diária"
    SEMANAL = "Semanal"
    MENSAL = "Mensal"

class TipoEscala(Enum):
    SEG_SEX = "5x2 (Segunda a Sexta)"
    SEXTA_SEGUNDA = "5x2 (Sexta a Terça)"
    ESCALA_6x1 = "6x1 (Domingo trabalhado)"
    ESCALA_12x36 = "12x36 (Plantão)"
    ESCALA_24x48 = "24x48 (Plantão longo)"
    PERSONALIZADA = "Personalizada"

class TipoSobreaviso(Enum):
    TODOS_DIAS = "Todos os dias selecionados"
    SEMANAS_ESPECIFICAS = "Só em semanas específicas"
    DIAS_UTEIS_FIXO = "Seg-Sex fixo, Sáb-Dom alternados"
    FINS_SEMANA_ALTERNADOS = "Sáb-Dom alternados"
    DIAS_UTEIS_ALTERNADOS = "Seg-Sex alternados"
    QUINZENAL = "Quinzenal (15 dias sim, 15 dias não)"

WEEKDAYS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
WD_MAP = {"Seg": 0, "Ter": 1, "Qua": 2, "Qui": 3, "Sex": 4, "Sáb": 5, "Dom": 6}
FREQ_OPTIONS = ["Toda semana", "Semanas pares", "Semanas ímpares", "Só em semanas de sobreaviso"]
WEEKEND_MODES = ["Não trabalha", "Alternados", "Quantidade por mês"]
WEEK_OF_MONTH_OPTIONS = ["1ª semana", "2ª semana", "3ª semana", "4ª semana", "5ª semana"]

# Constantes legais
FATOR_CONVERSAO = 1.142857  # 1/0.875 = 1.142857
INTERVALO_INTRAJORNADA = 1.0
INTERVALO_INTRAJORNADA_4A6H = 0.25
INTERJORNADA_MINIMA = 11.0
LIMITE_DIARIO_TRABALHO = 12.0
JORNADA_SEMANAL_PADRAO = 44.0

# Feriados nacionais brasileiros
FERIADOS_NACIONAIS = [
    {"nome": "Confraternização Universal", "data": "01/01"},
    {"nome": "Tiradentes", "data": "21/04"},
    {"nome": "Dia do Trabalho", "data": "01/05"},
    {"nome": "Independência do Brasil", "data": "07/09"},
    {"nome": "Nossa Sra. Aparecida", "data": "12/10"},
    {"nome": "Finados", "data": "02/11"},
    {"nome": "Proclamação da República", "data": "15/11"},
    {"nome": "Natal", "data": "25/12"},
]


# ----------------------------
# CLASSE PARA CONFIGURAÇÃO
# ----------------------------
class ConfiguracaoJornada:
    def __init__(self):
        self.tipo_apuracao = TipoApuracao.DIARIA
        self.jornada_diaria = 8.0
        self.jornada_semanal = 44.0
        self.jornada_mensal = 220.0
        self.limite_hora_extra = 8.0

        # Percentuais
        self.percentual_hora_extra = 0.50
        self.percentual_hora_extra_feriado = 1.00
        self.percentual_adicional_noturno = 0.20
        self.percentual_suminula_60 = 0.20

        # Escala
        self.tipo_escala = TipoEscala.SEG_SEX
        self.dias_escala_personalizada = []


# ----------------------------
# CLASSE HORARIO - CORAÇÃO DO SISTEMA
# ----------------------------
class Horario:
    """
    Classe que representa um horário de forma NÃO AMBÍGUA.
    REGRA ÚNICA: Se a hora for menor que a referência E for madrugada (0-11h), é do dia seguinte.
    """

    def __init__(self, hora_str: str, data_ref: date, hora_referencia: float = None):
        self.data_ref = data_ref
        self.hora_str = hora_str

        if not hora_str or pd.isna(hora_str) or hora_str == "":
            self.existe = False
            self.hora = 0
            self.minuto = 0
            self.valor_base = 0
            self.dias_adicionar = 0
            return

        try:
            if ":" in hora_str:
                h, m = map(int, hora_str.split(":"))
            else:
                h = int(hora_str)
                m = 0

            self.hora = h
            self.minuto = m
            self.existe = True
            self.valor_base = h + m/60.0

            # REGRA ÚNICA: Se a hora for menor que a referência E for madrugada (0-11h)
            if hora_referencia is not None:
                if self.valor_base < hora_referencia and self.valor_base < 12:
                    self.dias_adicionar = 1
                else:
                    self.dias_adicionar = 0
            else:
                self.dias_adicionar = 0

        except Exception as e:
            self.existe = False
            self.hora = 0
            self.minuto = 0
            self.valor_base = 0
            self.dias_adicionar = 0

    def para_decimal(self) -> float:
        """Retorna o valor em horas decimais (já considerando dia seguinte)"""
        if not self.existe:
            return 0.0
        return self.valor_base + (self.dias_adicionar * 24)

    def para_exibicao(self) -> str:
        """Retorna no formato HH:MM (sempre no formato normal)"""
        if not self.existe:
            return ""
        return f"{self.hora:02d}:{self.minuto:02d}"

    def __repr__(self):
        return f"Horario('{self.hora_str}', decimal={self.para_decimal():.2f})"


# ----------------------------
# CLASSE PERIODO - REPRESENTA UM TRECHO CONTÍNUO
# ----------------------------
class Periodo:
    def __init__(self, inicio: Horario, fim: Horario):
        self.inicio = inicio
        self.fim = fim

    @property
    def duracao(self) -> float:
        """Duração em horas decimais"""
        if not self.valido:
            return 0.0
        return self.fim.para_decimal() - self.inicio.para_decimal()

    @property
    def valido(self) -> bool:
        """Verifica se o período é válido"""
        return (self.inicio.existe and self.fim.existe and 
                self.fim.para_decimal() > self.inicio.para_decimal())

    def calcular_horas(self) -> Dict:
        """Calcula diurnas, noturnas e súmula 60 para este período"""
        if not self.valido:
            return {"diurnas": 0, "noturnas": 0, "sumula": 0, "total": 0}

        inicio_dec = self.inicio.para_decimal()
        fim_dec = self.fim.para_decimal()

        # Horas noturnas (22h às 29h)
        noturnas = 0.0
        hora_atual = inicio_dec
        while hora_atual < fim_dec:
            noite_inicio = max(hora_atual, 22.0)
            noite_fim = min(fim_dec, 29.0)
            if noite_fim > noite_inicio:
                noturnas += noite_fim - noite_inicio
            hora_atual += 24

        # Súmula 60 (após 29h)
        sumula = 0.0
        hora_atual = inicio_dec
        while hora_atual < fim_dec:
            sumula_inicio = max(hora_atual, 29.0)
            if sumula_inicio < fim_dec:
                sumula += fim_dec - sumula_inicio
            hora_atual += 24

        # Diurnas = total - noturnas - sumula
        diurnas = self.duracao - noturnas - sumula

        return {
            "diurnas": diurnas,
            "noturnas": noturnas,
            "sumula": sumula,
            "total": self.duracao
        }


# ----------------------------
# CLASSE JORNADA - REPRESENTA UM DIA INTEIRO
# ----------------------------
class Jornada:
    def __init__(self, data: date, C: str, D: str, E: str, F: str):
        self.data = data

        # Criar horários com referência correta em cadeia
        self.C = Horario(C, data)  # Início, sem referência

        # Para o fim do primeiro período, referência é o início
        self.D = Horario(D, data, self.C.para_decimal() if self.C.existe else None)

        # Para o início do segundo período, referência é o fim do primeiro
        if self.D.existe:
            self.E = Horario(E, data, self.D.para_decimal())
        else:
            self.E = Horario(E, data, self.C.para_decimal() if self.C.existe else None)

        # Para o fim da jornada, referência é o início do segundo período
        if self.E.existe:
            self.F = Horario(F, data, self.E.para_decimal())
        elif self.D.existe:
            self.F = Horario(F, data, self.D.para_decimal())
        else:
            self.F = Horario(F, data, self.C.para_decimal() if self.C.existe else None)

    def obter_periodos(self) -> List[Periodo]:
        """Retorna todos os períodos de trabalho (incluindo intervalo)"""
        periodos = []

        # Primeiro período
        if self.C.existe and self.D.existe:
            periodos.append(Periodo(self.C, self.D))

        # Intervalo (SEMPRE é hora trabalhada!)
        if self.D.existe and self.E.existe:
            periodos.append(Periodo(self.D, self.E))

        # Segundo período
        if self.E.existe and self.F.existe:
            periodos.append(Periodo(self.E, self.F))

        return [p for p in periodos if p.valido]

    def calcular(self) -> Dict:
        """Calcula todas as horas da jornada"""
        periodos = self.obter_periodos()

        total_diurnas = 0.0
        total_noturnas = 0.0
        total_sumula = 0.0
        total_bruto = 0.0

        for periodo in periodos:
            calc = periodo.calcular_horas()
            total_diurnas += calc["diurnas"]
            total_noturnas += calc["noturnas"]
            total_sumula += calc["sumula"]
            total_bruto += calc["total"]

        # Aplicar fator de redução noturna
        noturnas_convertidas = total_noturnas * FATOR_CONVERSAO
        sumula_convertidas = total_sumula * FATOR_CONVERSAO

        # Calcular horas extras
        horas_extras = max(0, (total_diurnas + noturnas_convertidas + sumula_convertidas) - 8.0)
        horas_normais = min(8.0, (total_diurnas + noturnas_convertidas + sumula_convertidas))

        return {
            "diurnas": round(total_diurnas, 2),
            "noturnas": round(noturnas_convertidas, 2),
            "sumula": round(sumula_convertidas, 2),
            "total": round(total_diurnas + noturnas_convertidas + sumula_convertidas, 2),
            "extras": round(horas_extras, 2),
            "normais": round(horas_normais, 2),
            "periodos": [(p.inicio.para_decimal(), p.fim.para_decimal()) for p in periodos],
            "periodos_str": " | ".join([f"{p.inicio.para_exibicao()}-{p.fim.para_exibicao()}" for p in periodos]),
            "fim_jornada": self.F.para_decimal() if self.F.existe else 0
        }


# ----------------------------
# FUNÇÕES HELPER
# ----------------------------
def is_missing_date(x) -> bool:
    if x is None:
        return True
    try:
        return bool(pd.isna(x))
    except Exception:
        return False

def as_date(x):
    if is_missing_date(x):
        return None
    if isinstance(x, pd.Timestamp):
        return x.date()
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None

def daterange(d0: date, d1: date):
    d = d0
    while d <= d1:
        yield d
        d += timedelta(days=1)

def weekday_str(d: date) -> str:
    for k, v in WD_MAP.items():
        if d.weekday() == v:
            return k
    return "?"

def is_weekend(d: date) -> bool:
    return d.weekday() in (5, 6)

def week_start_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def week_of_month(d: date) -> int:
    first_day = date(d.year, d.month, 1)
    first_week_start = first_day - timedelta(days=first_day.weekday())
    current_week_start = d - timedelta(days=d.weekday())
    week_number = ((current_week_start - first_week_start).days // 7) + 1
    return min(week_number, 5)

def in_any_period(d: date, periods):
    for p in periods:
        s = p.get("start")
        e = p.get("end")
        if s is None or e is None:
            continue
        if s <= d <= e:
            return True
    return False

def pick_month_weekends(start_date: date, end_date: date, sat_per_month: int, sun_per_month: int):
    chosen = set()
    cur = date(start_date.year, start_date.month, 1)

    while cur <= end_date:
        if cur.month == 12:
            last = date(cur.year, 12, 31)
        else:
            last = date(cur.year, cur.month + 1, 1) - timedelta(days=1)

        m_start = max(cur, start_date)
        m_end = min(last, end_date)

        sats = []
        suns = []
        d = m_start
        while d <= m_end:
            if d.weekday() == 5:
                sats.append(d)
            elif d.weekday() == 6:
                suns.append(d)
            d += timedelta(days=1)

        for dd in sats[:sat_per_month]:
            chosen.add(dd)
        for dd in suns[:sun_per_month]:
            chosen.add(dd)

        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)

    return chosen

def gerar_escala_12x36(start_date: date, end_date: date) -> List[date]:
    dias_trabalho = []
    trabalha = True
    for d in daterange(start_date, end_date):
        if trabalha:
            dias_trabalho.append(d)
        trabalha = not trabalha
    return dias_trabalho

def gerar_escala_24x48(start_date: date, end_date: date) -> List[date]:
    dias_trabalho = []
    dia_count = 0
    for d in daterange(start_date, end_date):
        if dia_count % 3 == 0:
            dias_trabalho.append(d)
        dia_count += 1
    return dias_trabalho

def gerar_escala_6x1(start_date: date, end_date: date) -> List[date]:
    dias_trabalho = []
    for d in daterange(start_date, end_date):
        if d.weekday() != 6:
            dias_trabalho.append(d)
    return dias_trabalho

def get_feriados_no_periodo(start_date: date, end_date: date, feriados_selecionados: list) -> List[date]:
    feriados = []
    for feriado in feriados_selecionados:
        data_feriado = datetime.strptime(feriado["data"] + f"/{start_date.year}", "%d/%m/%Y").date()
        if start_date <= data_feriado <= end_date:
            feriados.append(data_feriado)
    return feriados

def parse_periods_from_text(text: str):
    periods = []
    lines = text.strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue
        line = ' '.join(line.split())

        patterns = [
            r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(.+)',
            r'(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})\s+(.+)',
            r'(\d{2}/\d{2}/\d{4})-(\d{2}/\d{2}/\d{4})\s+(.+)',
            r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})',
        ]

        match = None
        for pattern in patterns:
            match = re.match(pattern, line)
            if match:
                break

        if match:
            groups = match.groups()
            if len(groups) >= 2:
                try:
                    start_date = datetime.strptime(groups[0], "%d/%m/%Y").date()
                    end_date = datetime.strptime(groups[1], "%d/%m/%Y").date()
                    tipo = groups[2] if len(groups) >= 3 else "Férias"
                    periods.append({
                        "Data Inicial": start_date,
                        "Data Final": end_date,
                        "Tipo": tipo.strip()
                    })
                except ValueError:
                    pass

    return pd.DataFrame(periods) if periods else pd.DataFrame(columns=["Data Inicial", "Data Final", "Tipo"])


# ----------------------------
# FUNÇÃO PARA CÁLCULO DE SOBREAVISO (CORRIGIDA)
# ----------------------------
def calcular_sobreaviso(fim_hoje: float, inicio_amanha: float, acion_horas: float) -> Tuple[float, float]:
    """
    Calcula o sobreaviso de forma robusta, SEMPRE positivo
    """
    if fim_hoje <= 0 or inicio_amanha <= 0:
        return 0.0, 0.0

    # Caso 1: Jornada terminou no dia seguinte (fim_hoje > 24)
    # e amanhã começa cedo (inicio_amanha < 12)
    if fim_hoje > 24 and inicio_amanha < 12:
        inicio_amanha += 24

    # Caso 2: Jornada terminou no mesmo dia (fim_hoje < 24)
    # e amanhã começa cedo (inicio_amanha < 12)
    elif fim_hoje < 24 and inicio_amanha < 12:
        inicio_amanha += 24

    # Calcular diferença
    so_bruto = inicio_amanha - fim_hoje

    # Garantir que seja positivo
    if so_bruto < 0:
        so_bruto += 24

    # Limitar a no máximo 24h
    so_bruto = min(so_bruto, 24.0)

    # Sobreaviso líquido = bruto - acionamentos
    so_liq = max(0.0, so_bruto - acion_horas)

    return round(so_bruto, 2), round(so_liq, 2)


# ----------------------------
# FUNÇÃO PARA CÁLCULO DE INTERJORNADA - CORRETA E DEFINITIVA
# ----------------------------
def calcular_interjornada(fim_hoje: float, inicio_amanha: float) -> float:
    """
    Calcula o DÉFICIT de interjornada.
    Retorna quantas horas faltam para completar 11h de descanso.
    Se descansou 11h ou mais, retorna 0.
    """
    if fim_hoje <= 0 or inicio_amanha <= 0:
        return 0.0

    # Calcular intervalo real entre o fim de hoje e início de amanhã
    intervalo = inicio_amanha - fim_hoje

    # Ajustar se passou da meia-noite
    if intervalo < 0:
        intervalo += 24

    intervalo = round(max(0.0, intervalo), 2)

    # REGRA: Se descansou MENOS que 11h, retorna o DÉFICIT (11 - intervalo)
    # Se descansou 11h ou MAIS, retorna 0
    if intervalo < INTERJORNADA_MINIMA:
        return round(INTERJORNADA_MINIMA - intervalo, 2)
    else:
        return 0.0


# ----------------------------
# FUNÇÃO PARA CÁLCULO DE INTRAJORNADA
# ----------------------------
def calcular_intrajornada(periodos: List) -> float:
    """
    Calcula o intervalo intrajornada (soma dos intervalos entre períodos)
    Considera que o intervalo é o período entre D e E
    """
    if len(periodos) < 2:
        return 0.0

    # O intervalo intrajornada é o segundo período (entre D e E)
    if len(periodos) >= 2:
        return round(periodos[1].duracao, 2) if periodos[1].valido else 0.0

    return 0.0


# ----------------------------
# FUNÇÃO PRINCIPAL
# ----------------------------
def build_card(
    start_date: date,
    end_date: date,
    base_days_schedule: dict,
    weekend_mode: str,
    weekend_anchor: date,
    sat_per_month: int,
    sun_per_month: int,
    sobreaviso_days: set,
    sobreaviso_mode: str,
    so_fixed_start: str,
    so_fixed_end: str,
    acionamentos_rules_df: pd.DataFrame,
    off_periods_df: pd.DataFrame,
    feriados_trabalhados: List[date],
    config: ConfiguracaoJornada,
    tipo_sobreaviso: TipoSobreaviso = TipoSobreaviso.TODOS_DIAS,
    semanas_sobreaviso: list = None,
    sobreaviso_alternado_anchor: date = None,
):
    if semanas_sobreaviso is None:
        semanas_sobreaviso = []

    if sobreaviso_alternado_anchor is None:
        sobreaviso_alternado_anchor = start_date

    dias_trabalho_escala = []
    if config.tipo_escala == TipoEscala.ESCALA_12x36:
        dias_trabalho_escala = gerar_escala_12x36(start_date, end_date)
    elif config.tipo_escala == TipoEscala.ESCALA_24x48:
        dias_trabalho_escala = gerar_escala_24x48(start_date, end_date)
    elif config.tipo_escala == TipoEscala.ESCALA_6x1:
        dias_trabalho_escala = gerar_escala_6x1(start_date, end_date)

    off_periods = []
    for _, r in off_periods_df.iterrows():
        s = as_date(r.get("Data Inicial"))
        e = as_date(r.get("Data Final"))
        if s is None or e is None:
            continue
        off_periods.append({"start": s, "end": e, "tipo": str(r.get("Tipo", "")).strip()})

    rules = []
    for _, r in acionamentos_rules_df.iterrows():
        ativo = r.get("Ativo", True)
        if pd.isna(ativo):
            ativo = True
        if not bool(ativo):
            continue

        wd = str(r.get("Dia", "")).strip()
        if wd not in WD_MAP:
            continue

        hs = str(r.get("G", "")).strip()
        he = str(r.get("H", "")).strip()
        if not hs or not he:
            continue

        freq = str(r.get("Freq", "Toda semana")).strip()
        if freq not in FREQ_OPTIONS:
            freq = "Toda semana"

        v_ini = as_date(r.get("Vigência Início"))
        v_fim = as_date(r.get("Vigência Fim"))

        rules.append({
            "wd": wd,
            "start": hs,
            "end": he,
            "freq": freq,
            "v_ini": v_ini,
            "v_fim": v_fim,
        })

    weekend_work_dates = set()
    if weekend_mode == "Quantidade por mês":
        weekend_work_dates = pick_month_weekends(start_date, end_date, int(sat_per_month), int(sun_per_month))

    def compute_has_base(d: date) -> bool:
        wd = weekday_str(d)

        # Se for feriado, só trabalha se estiver na lista de feriados trabalhados
        for feriado in FERIADOS_NACIONAIS:
            data_feriado = datetime.strptime(feriado["data"] + f"/{d.year}", "%d/%m/%Y").date()
            if data_feriado == d:
                # É feriado, só trabalha se estiver na lista de trabalhados
                return d in feriados_trabalhados

        if in_any_period(d, off_periods):
            return False

        if dias_trabalho_escala and d not in dias_trabalho_escala:
            return False

        has = (wd in base_days_schedule)

        if is_weekend(d):
            if weekend_mode == "Não trabalha":
                has = False
            elif weekend_mode == "Alternados":
                ws_anchor = week_start_monday(weekend_anchor)
                ws_cur = week_start_monday(d)
                diff_weeks = (ws_cur - ws_anchor).days // 7
                has = (diff_weeks % 2 == 0)
            elif weekend_mode == "Quantidade por mês":
                has = (d in weekend_work_dates)

        return has

    def get_horarios_dia(d: date):
        wd = weekday_str(d)
        if wd in base_days_schedule:
            return base_days_schedule[wd]
        return ("", "", "", "")

    def tem_sobreaviso(d: date) -> bool:
        wd = weekday_str(d)
        if wd not in sobreaviso_days:
            return False
        if tipo_sobreaviso == TipoSobreaviso.TODOS_DIAS:
            return True
        elif tipo_sobreaviso == TipoSobreaviso.SEMANAS_ESPECIFICAS:
            if not semanas_sobreaviso:
                return True
            semana_mes = week_of_month(d)
            return semana_mes in semanas_sobreaviso
        elif tipo_sobreaviso == TipoSobreaviso.DIAS_UTEIS_FIXO:
            if wd in ["Seg", "Ter", "Qua", "Qui", "Sex"]:
                return True
            elif wd in ["Sáb", "Dom"]:
                ws_anchor = week_start_monday(sobreaviso_alternado_anchor)
                ws_cur = week_start_monday(d)
                diff_weeks = (ws_cur - ws_anchor).days // 7
                return (diff_weeks % 2 == 0)
        elif tipo_sobreaviso == TipoSobreaviso.FINS_SEMANA_ALTERNADOS:
            if wd in ["Sáb", "Dom"]:
                ws_anchor = week_start_monday(sobreaviso_alternado_anchor)
                ws_cur = week_start_monday(d)
                diff_weeks = (ws_cur - ws_anchor).days // 7
                return (diff_weeks % 2 == 0)
            else:
                return False
        elif tipo_sobreaviso == TipoSobreaviso.DIAS_UTEIS_ALTERNADOS:
            if wd in ["Seg", "Ter", "Qua", "Qui", "Sex"]:
                ws_anchor = week_start_monday(sobreaviso_alternado_anchor)
                ws_cur = week_start_monday(d)
                diff_weeks = (ws_cur - ws_anchor).days // 7
                return (diff_weeks % 2 == 0)
            else:
                return False
        elif tipo_sobreaviso == TipoSobreaviso.QUINZENAL:
            dias_desde_anchor = (d - sobreaviso_alternado_anchor).days
            quinzena = (dias_desde_anchor // 15) % 2
            return quinzena == 0
        return True

    semana_atual = None
    horas_acumuladas_semana = 0
    horas_acumuladas_mes = 0

    rows = []
    irregularidades = []

    # Dicionário para armazenar jornadas anteriores
    jornadas_anteriores = {}

    for d in daterange(start_date, end_date):
        semana_dia = d.isocalendar()[1]
        if semana_atual != semana_dia:
            semana_atual = semana_dia
            horas_acumuladas_semana = 0

        wd = weekday_str(d)
        is_off = in_any_period(d, off_periods)
        is_weekend_day = is_weekend(d)
        is_feriado = d in feriados_trabalhados
        is_domingo_ou_feriado = (wd == "Dom") or is_feriado

        # Horários base
        C = D = E = F = ""
        G = H = ""  # Acionamentos

        has_base_today = compute_has_base(d)
        if has_base_today and not is_off:
            C, D, E, F = get_horarios_dia(d)

        # ============= NOVO SISTEMA - JORNADA =============
        jornada = Jornada(d, C, D, E, F)
        calculo = jornada.calcular()

        # Períodos em string para debug
        periodos_str = " | ".join([f"{p[0]:.1f}-{p[1]:.1f}" for p in calculo["periodos"]])

        # ============= ACIONAMENTOS =============
        acion_horas = 0.0
        acion_list = []
        if not is_off:
            for rule in rules:
                if rule["wd"] != wd:
                    continue
                if rule["v_ini"] is not None and d < rule["v_ini"]:
                    continue
                if rule["v_fim"] is not None and d > rule["v_fim"]:
                    continue

                w = d.isocalendar().week

                if rule["freq"] == "Só em semanas de sobreaviso":
                    if not tem_sobreaviso(d):
                        continue
                elif rule["freq"] == "Semanas pares" and (w % 2 != 0):
                    continue
                elif rule["freq"] == "Semanas ímpares" and (w % 2 == 0):
                    continue

                if not G:
                    G, H = rule["start"], rule["end"]

                # Processa acionamento
                try:
                    h_g, m_g = map(int, rule["start"].split(":"))
                    h_h, m_h = map(int, rule["end"].split(":"))

                    acion_inicio = Horario(rule["start"], d)
                    acion_fim = Horario(rule["end"], d, acion_inicio.para_decimal())

                    if acion_fim.existe and acion_inicio.existe:
                        horas_acion = acion_fim.para_decimal() - acion_inicio.para_decimal()
                        if horas_acion > 0:
                            acion_horas += horas_acion
                            acion_list.append(f"{rule['start']}-{rule['end']}")
                except:
                    pass

        # ============= SOBREAVISO CORRIGIDO =============
        so_bruto = 0.0
        so_liq = 0.0

        if tem_sobreaviso(d) and (not is_off):
            if sobreaviso_mode == "Fim da jornada → Próx. entrada":
                next_day = d + timedelta(days=1)
                if next_day <= end_date:
                    has_base_next = compute_has_base(next_day)
                    if has_base_next:
                        next_C, _, _, _ = get_horarios_dia(next_day)
                        if next_C:
                            # Fim da jornada de hoje
                            fim_hoje = jornada.F.para_decimal() if jornada.F.existe else 0
                            # Início da jornada de amanhã
                            inicio_amanha = Horario(next_C, next_day).para_decimal()

                            so_bruto, so_liq = calcular_sobreaviso(fim_hoje, inicio_amanha, acion_horas)
            else:  # Fixo
                try:
                    so_inicio = Horario(so_fixed_start, d)
                    so_fim = Horario(so_fixed_end, d, so_inicio.para_decimal())

                    if so_fim.existe and so_inicio.existe:
                        so_bruto = so_fim.para_decimal() - so_inicio.para_decimal()
                        if so_bruto < 0:
                            so_bruto += 24
                        so_bruto = min(so_bruto, 24.0)
                        so_liq = max(0.0, so_bruto - acion_horas)
                except:
                    pass

        # ============= CÁLCULO DE INTERJORNADA E INTRAJORNADA =============
        intrajornada = calcular_intrajornada(jornada.obter_periodos())

        # ============= CÁLCULO DA INTERJORNADA (COM DESCONTO DOS ACIONAMENTOS) =============
        interjornada_intervalo = 0.0
        horas_inter_devidas = 0.0

        # Só calcula interjornada se o dia TEM jornada
        if jornada.C.existe:
            # Procurar o último dia que teve jornada antes deste dia
            ultimo_dia_trabalhado = None

            for dia_temp in range(1, 30):
                dia_anterior_busca = d - timedelta(days=dia_temp)
                if dia_anterior_busca < start_date:
                    break
                if dia_anterior_busca in jornadas_anteriores:
                    jornada_anterior = jornadas_anteriores[dia_anterior_busca]
                    if jornada_anterior.F.existe:
                        ultimo_dia_trabalhado = dia_anterior_busca
                        break

            if ultimo_dia_trabalhado is not None:
                jornada_anterior = jornadas_anteriores[ultimo_dia_trabalhado]
                fim_anterior = jornada_anterior.F.para_decimal()
                inicio_atual = jornada.C.para_decimal()

                # Calcular intervalo REAL entre o último dia trabalhado e hoje
                interjornada_intervalo = inicio_atual - fim_anterior

                # Se o início atual for menor que o fim anterior, significa que passou da meia-noite
                if interjornada_intervalo < 0:
                    interjornada_intervalo += 24

                interjornada_intervalo = round(max(0.0, interjornada_intervalo), 2)

                # IMPORTANTE: Descontar as horas de acionamento que ocorreram durante o descanso
                # As horas de acionamento já foram calculadas no dia anterior (acion_horas)
                # E representam trabalho, não descanso!
                descanso_real = max(0.0, interjornada_intervalo - acion_horas)

                # Calcular horas devidas (déficit) baseado no descanso REAL
                if descanso_real < INTERJORNADA_MINIMA:
                    horas_inter_devidas = round(INTERJORNADA_MINIMA - descanso_real, 2)

                    irregularidades.append({
                        "data": d,
                        "tipo": "Intervalo interjornada insuficiente",
                        "detalhe": f"Descanso bruto: {interjornada_intervalo:.2f}h, Acionamentos: {acion_horas:.2f}h, Descanso real: {descanso_real:.2f}h (mínimo: {INTERJORNADA_MINIMA}h) - Déficit: {horas_inter_devidas:.2f}h"
                    })

        # Armazenar jornada atual para uso no próximo dia
        jornadas_anteriores[d] = jornada

        # ============= NOVOS CÁLCULOS =============
        # Total Horas Líquidas = Total Horas - Intrajornada (APENAS para dias normais, excluindo domingos/feriados)
        if is_domingo_ou_feriado:
            # Se for domingo/feriado, não conta no total de horas líquidas
            total_horas_liquidas = 0.0
        else:
            # Se for dia normal, calcula normalmente
            total_horas_liquidas = calculo["total"] - intrajornada

        # Horas em Domingos/Feriados (usando total líquido, descontando intervalo)
        horas_dom_fer = 0.0
        horas_dom_fer_extra = 0.0

        if is_domingo_ou_feriado and has_base_today and not is_off:
            horas_dom_fer = calculo["total"] - intrajornada  # Coluna W - usa total líquido
            # Horas extras em domingos/feriados = horas que ultrapassam a jornada normal
            horas_dom_fer_extra = max(0, (calculo["total"] - intrajornada) - config.jornada_diaria)  # Coluna X - usa total líquido

        # Cálculo de horas extras (excluindo domingos/feriados)
        if is_domingo_ou_feriado:
            # Se for domingo/feriado, não calcula hora extra na coluna Q
            horas_extras = 0.0
        else:
            # Se não for domingo/feriado, calcula hora extra normalmente
            horas_extras = calculo["extras"]

        # Percentual de hora extra
        percentual_extra = config.percentual_hora_extra_feriado if is_feriado else config.percentual_hora_extra

        # Verificação de limites legais
        if calculo["total"] > LIMITE_DIARIO_TRABALHO:
            irregularidades.append({
                "data": d,
                "tipo": "Limite diário excedido",
                "detalhe": f"{calculo['total']:.2f}h (limite: {LIMITE_DIARIO_TRABALHO}h)"
            })

        # Montagem do DataFrame - REMOVIDA a coluna "Interjornada Intervalo (h)"
        rows.append({
            "Data": d.strftime("%d/%m/%Y"),
            "Dia Semana": wd,
            "Feriado": "Sim" if is_feriado else "Não",
            "C (Início)": C,
            "D (Saída)": D,
            "E (Retorno)": E,
            "F (Fim)": F,
            "G (Acionamento)": G,
            "H (Acionamento)": H,
            "Horas Diurnas (h)": calculo["diurnas"],
            "Horas Noturnas 22-5h (h)": calculo["noturnas"],
            "Horas Súmula 60 (h)": calculo["sumula"],
            "Total Horas (h)": calculo["total"],
            "Intrajornada (h)": intrajornada,
            "Total Horas Líquidas (h)": round(total_horas_liquidas, 2),
            "Horas Acionamento (h)": round(acion_horas, 2),
            "Horas Extras (h)": round(horas_extras, 2),
            "% Hora Extra": f"{percentual_extra*100:.0f}%",
            "Sobreaviso Bruto (h)": round(so_bruto, 2),
            "Sobreaviso Líquido (h)": round(so_liq, 2),
            "Sobreaviso Ativo": "Sim" if so_bruto > 0 else "Não",
            "Horas Normais (h)": calculo["normais"],
            "Horas Dom/Fer (h)": round(horas_dom_fer, 2),
            "Horas Dom/Fer Extras (h)": round(horas_dom_fer_extra, 2),
            "Interjornada (h)": horas_inter_devidas,  # APENAS O DÉFICIT (0 quando >= 11h)
            "Períodos (decimal)": periodos_str
        })

    df = pd.DataFrame(rows)
    df["Mês"] = pd.to_datetime(df["Data"], dayfirst=True).dt.to_period("M").astype(str)

    # Totais por mês - REMOVIDA "Interjornada Intervalo (h)"
    totais_por_mes = df.groupby("Mês", as_index=False).agg({
        "Horas Extras (h)": "sum",
        "Horas Diurnas (h)": "sum",
        "Horas Noturnas 22-5h (h)": "sum",
        "Horas Súmula 60 (h)": "sum",
        "Horas Acionamento (h)": "sum",
        "Sobreaviso Líquido (h)": "sum",
        "Total Horas (h)": "sum",
        "Total Horas Líquidas (h)": "sum",
        "Horas Normais (h)": "sum",
        "Intrajornada (h)": "sum",
        "Horas Dom/Fer (h)": "sum",
        "Horas Dom/Fer Extras (h)": "sum",
        "Interjornada (h)": "sum"  # Soma das horas devidas
    }).round(2)

    total_geral = {
        "Mês": "TOTAL GERAL",
        "Horas Extras (h)": df["Horas Extras (h)"].sum(),
        "Horas Diurnas (h)": df["Horas Diurnas (h)"].sum(),
        "Horas Noturnas 22-5h (h)": df["Horas Noturnas 22-5h (h)"].sum(),
        "Horas Súmula 60 (h)": df["Horas Súmula 60 (h)"].sum(),
        "Horas Acionamento (h)": df["Horas Acionamento (h)"].sum(),
        "Sobreaviso Líquido (h)": df["Sobreaviso Líquido (h)"].sum(),
        "Total Horas (h)": df["Total Horas (h)"].sum(),
        "Total Horas Líquidas (h)": df["Total Horas Líquidas (h)"].sum(),
        "Horas Normais (h)": df["Horas Normais (h)"].sum(),
        "Intrajornada (h)": df["Intrajornada (h)"].sum(),
        "Horas Dom/Fer (h)": df["Horas Dom/Fer (h)"].sum(),
        "Horas Dom/Fer Extras (h)": df["Horas Dom/Fer Extras (h)"].sum(),
        "Interjornada (h)": df["Interjornada (h)"].sum()
    }

    totais_por_mes = pd.concat([totais_por_mes, pd.DataFrame([total_geral])], ignore_index=True)

    df_irregularidades = pd.DataFrame(irregularidades)
    df_export = df.drop(columns=["Mês", "Períodos (decimal)"])

    return df_export, totais_por_mes, df_irregularidades


def dataframe_to_excel_bytes(df_card: pd.DataFrame, 
                            df_totais_mes: pd.DataFrame,
                            df_irregularidades: pd.DataFrame,
                            config: ConfiguracaoJornada,
                            tipo_sobreaviso: TipoSobreaviso,
                            semanas_sobreaviso: list = None,
                            sobreaviso_alternado_anchor: date = None) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Cartao_Ponto"
    for r in dataframe_to_rows(df_card, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="TabelaCartao", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    ws3 = wb.create_sheet("Totais")
    for r in dataframe_to_rows(df_totais_mes, index=False, header=True):
        ws3.append(r)

    for c in ws3[1]:
        c.fill = PatternFill("solid", fgColor="4F81BD")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for row in ws3.iter_rows(min_row=ws3.max_row, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
        for c in row:
            c.fill = PatternFill("solid", fgColor="FFC000")
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    ws3.freeze_panes = "A2"

    if not df_irregularidades.empty:
        ws4 = wb.create_sheet("Irregularidades")
        ws4.append(["Data", "Tipo", "Detalhe"])

        for _, row in df_irregularidades.iterrows():
            ws4.append([
                row["data"].strftime("%d/%m/%Y") if hasattr(row["data"], 'strftime') else str(row["data"]),
                row["tipo"],
                row["detalhe"]
            ])

        for c in ws4[1]:
            c.fill = PatternFill("solid", fgColor="C00000")
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# INTERFACE STREAMLIT
# ----------------------------
def render():

    st.title("📋 Cartão de Ponto - SISTEMA DEFINITIVO")
    st.caption("Funciona para QUALQUER horário: 11:00-23:00, 08:00-22:00, 16:00-04:00, 18:00-09:00, etc.")

    if 'config' not in st.session_state:
        st.session_state.config = ConfiguracaoJornada()

    with st.sidebar:
        st.header("⚙️ Configurações")

        with st.expander("📊 Apuração e Limites", expanded=True):
            st.session_state.config.tipo_apuracao = TipoApuracao(
                st.selectbox("Forma de apuração:", 
                            [t.value for t in TipoApuracao],
                            index=0)
            )

            col_lim1, col_lim2 = st.columns(2)
            with col_lim1:
                st.session_state.config.jornada_diaria = st.number_input(
                    "Jornada diária (h):", 
                    min_value=1.0, max_value=16.0, value=8.0, step=0.5
                )
            with col_lim2:
                st.session_state.config.limite_hora_extra = st.number_input(
                    "Limite p/ hora extra (h):", 
                    min_value=0.0, max_value=24.0, value=8.0, step=0.5
                )

            st.session_state.config.jornada_semanal = st.number_input(
                "Jornada semanal (h):", 
                min_value=30.0, max_value=60.0, value=44.0, step=0.5
            )
            st.session_state.config.jornada_mensal = st.number_input(
                "Jornada mensal (h):", 
                min_value=120.0, max_value=300.0, value=220.0, step=1.0
            )

        with st.expander("💰 Percentuais"):
            col_per1, col_per2, col_per3 = st.columns(3)
            with col_per1:
                st.session_state.config.percentual_hora_extra = st.number_input(
                    "% HE:", min_value=0, max_value=200, value=50
                ) / 100
            with col_per2:
                st.session_state.config.percentual_hora_extra_feriado = st.number_input(
                    "% HE Feriado:", min_value=0, max_value=200, value=100
                ) / 100
            with col_per3:
                st.session_state.config.percentual_adicional_noturno = st.number_input(
                    "% Adic. Noturno:", min_value=0, max_value=100, value=20
                ) / 100

            st.session_state.config.percentual_suminula_60 = st.number_input(
                "% Súmula 60:", min_value=0, max_value=100, value=20
            ) / 100

        with st.expander("📅 Escala"):
            st.session_state.config.tipo_escala = TipoEscala(
                st.selectbox("Escala:", 
                            [t.value for t in TipoEscala],
                            index=0)
            )

    # Data range
    colA, colB = st.columns(2)
    with colA:
        start_date = st.date_input("Data inicial", value=date(2024, 1, 1), format="DD/MM/YYYY")
    with colB:
        end_date = st.date_input("Data final", value=date(2024, 12, 31), format="DD/MM/YYYY")

    st.divider()

    # ============= JORNADA BASE =============
    st.subheader("1) Jornada Base")
    st.caption("C = Início | D = Saída p/ intervalo | E = Retorno | F = Fim")
    st.info("💡 O sistema funciona para QUALQUER horário. Exemplos: 11:00-23:00, 08:00-22:00, 16:00-04:00, 18:00-09:00")

    # Templates rápidos para teste
    st.write("### Templates para teste:")
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        if st.button("🌞 08-22"):
            st.session_state.template = ("08:00", "12:00", "13:00", "22:00")
    with col2:
        if st.button("🌙 18-09"):
            st.session_state.template = ("18:00", "02:00", "03:00", "09:00")
    with col3:
        if st.button("🌙 20-08"):
            st.session_state.template = ("20:00", "00:00", "01:00", "08:00")
    with col4:
        if st.button("🏥 16-04"):
            st.session_state.template = ("16:00", "20:00", "21:00", "04:00")
    with col5:
        if st.button("📋 11-23"):
            st.session_state.template = ("11:00", "16:00", "17:00", "23:00")

    # Criar DataFrame para edição
    jornada_data = []
    for day in WEEKDAYS:
        if hasattr(st.session_state, 'template'):
            c, d, e, f = st.session_state.template
            trabalha = True
        else:
            # Valores padrão
            if day in ["Seg", "Ter", "Qua", "Qui", "Sex"]:
                c, d, e, f = "08:00", "12:00", "13:00", "17:00"
                trabalha = True
            else:
                c, d, e, f = "", "", "", ""
                trabalha = False

        jornada_data.append({
            "Dia": day,
            "Trabalha": trabalha,
            "C (Início)": c,
            "D (Saída)": d,
            "E (Retorno)": e,
            "F (Fim)": f
        })

    jornada_df = pd.DataFrame(jornada_data)

    edited_jornada = st.data_editor(
        jornada_df,
        column_config={
            "Dia": st.column_config.TextColumn("Dia", disabled=True, width="small"),
            "Trabalha": st.column_config.CheckboxColumn("Trab", width="small"),
            "C (Início)": st.column_config.TextColumn("C", width="small", max_chars=5),
            "D (Saída)": st.column_config.TextColumn("D", width="small", max_chars=5),
            "E (Retorno)": st.column_config.TextColumn("E", width="small", max_chars=5),
            "F (Fim)": st.column_config.TextColumn("F", width="small", max_chars=5),
        },
        use_container_width=True,
        hide_index=True,
        num_rows="fixed"
    )

    base_days_schedule = {}
    for _, row in edited_jornada.iterrows():
        if row["Trabalha"]:
            base_days_schedule[row["Dia"]] = (
                row["C (Início)"],
                row["D (Saída)"],
                row["E (Retorno)"],
                row["F (Fim)"]
            )

    st.divider()

    # ============= FERIADOS =============
    st.subheader("2) Feriados")

    col_feriado1, col_feriado2 = st.columns([1, 3])
    with col_feriado1:
        tem_feriados = st.checkbox("Houve feriados trabalhados?", value=False)

    feriados_trabalhados = []
    if tem_feriados:
        st.caption("Selecione os feriados que foram trabalhados:")

        feriados_opcoes = []
        for feriado in FERIADOS_NACIONAIS:
            data_feriado = datetime.strptime(feriado["data"] + f"/{start_date.year}", "%d/%m/%Y").date()
            if start_date <= data_feriado <= end_date:
                feriados_opcoes.append({
                    "nome": feriado["nome"],
                    "data": data_feriado,
                    "label": f"{data_feriado.strftime('%d/%m')} - {feriado['nome']}"
                })

        if feriados_opcoes:
            feriados_selecionados = st.multiselect(
                "Feriados:",
                options=[f["label"] for f in feriados_opcoes],
                default=[]
            )

            feriados_trabalhados = [
                f["data"] for f in feriados_opcoes 
                if f["label"] in feriados_selecionados
            ]
        else:
            st.info("Não há feriados nacionais no período selecionado.")
    else:
        # Se não marcou que houve feriados trabalhados, a lista fica vazia
        feriados_trabalhados = []

    st.divider()

    # ============= FINS DE SEMANA =============
    st.subheader("3) Fins de semana")
    weekend_mode = st.radio("", WEEKEND_MODES, horizontal=True, label_visibility="collapsed")

    weekend_anchor = start_date
    sat_per_month = 0
    sun_per_month = 0

    if weekend_mode == "Alternados":
        weekend_anchor = st.date_input("Data âncora", value=start_date, format="DD/MM/YYYY")

    if weekend_mode == "Quantidade por mês":
        c1, c2 = st.columns(2)
        with c1:
            sat_per_month = st.number_input("Sábados/mês", 0, 5, 2, 1)
        with c2:
            sun_per_month = st.number_input("Domingos/mês", 0, 5, 1, 1)

    st.divider()

    # ============= SOBREAVISO =============
    st.subheader("4) Sobreaviso")

    sobreaviso_days = set(st.multiselect("Dias com sobreaviso:", WEEKDAYS, default=WEEKDAYS))

    tipos_sobreaviso = [t.value for t in TipoSobreaviso]
    tipo_sobreaviso_selecionado = st.selectbox("Padrão:", tipos_sobreaviso, index=0)
    tipo_sobreaviso = TipoSobreaviso(tipo_sobreaviso_selecionado)

    sobreaviso_alternado_anchor = None
    semanas_sobreaviso = []

    if tipo_sobreaviso == TipoSobreaviso.SEMANAS_ESPECIFICAS:
        semanas_sobreaviso = st.multiselect("Semanas:", WEEK_OF_MONTH_OPTIONS, default=["1ª semana", "3ª semana"])
        semanas_numeros = []
        for s in semanas_sobreaviso:
            if s == "1ª semana": semanas_numeros.append(1)
            elif s == "2ª semana": semanas_numeros.append(2)
            elif s == "3ª semana": semanas_numeros.append(3)
            elif s == "4ª semana": semanas_numeros.append(4)
            elif s == "5ª semana": semanas_numeros.append(5)
        semanas_sobreaviso = semanas_numeros

    elif tipo_sobreaviso in [TipoSobreaviso.DIAS_UTEIS_FIXO, 
                             TipoSobreaviso.FINS_SEMANA_ALTERNADOS,
                             TipoSobreaviso.DIAS_UTEIS_ALTERNADOS,
                             TipoSobreaviso.QUINZENAL]:
        sobreaviso_alternado_anchor = st.date_input("Âncora:", value=start_date, format="DD/MM/YYYY")

    sobreaviso_mode = st.radio("Modo:", ["Fim da jornada", "Fixo"], horizontal=True)
    so_fixed_start = "19:30"
    so_fixed_end = "08:00"

    if sobreaviso_mode == "Fixo":
        c1, c2 = st.columns(2)
        with c1: so_fixed_start = st.text_input("Início", value="19:30", max_chars=5)
        with c2: so_fixed_end = st.text_input("Fim", value="08:00", max_chars=5)

    st.divider()

    # ============= ACIONAMENTOS =============
    st.subheader("5) Acionamentos (G, H)")
    default_rules = pd.DataFrame([
        {"Ativo": True, "Dia": "Ter", "G": "11:00", "H": "12:00", "Freq": "Toda semana"},
        {"Ativo": True, "Dia": "Qui", "G": "11:00", "H": "12:00", "Freq": "Toda semana"},
        {"Ativo": True, "Dia": "Sex", "G": "11:00", "H": "12:00", "Freq": "Toda semana"},
    ])

    rules_df = st.data_editor(
        default_rules,
        num_rows="dynamic",
        column_config={
            "Ativo": st.column_config.CheckboxColumn("Ativo", width="small"),
            "Dia": st.column_config.SelectboxColumn("Dia", options=WEEKDAYS, width="small"),
            "G": st.column_config.TextColumn("G (Início)", width="small", max_chars=5),
            "H": st.column_config.TextColumn("H (Fim)", width="small", max_chars=5),
            "Freq": st.column_config.SelectboxColumn("Frequência", options=FREQ_OPTIONS, width="medium"),
        },
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    # ============= FÉRIAS / AFASTAMENTOS =============
    st.subheader("6) Férias / Afastamentos")
    input_method = st.radio("", ["Colar", "Manual"], horizontal=True)

    if input_method == "Colar":
        text_input = st.text_area("", height=100, placeholder="Ex: 01/01/2024 a 15/01/2024 Férias")
        if st.button("Processar", use_container_width=True):
            off_df = parse_periods_from_text(text_input)
        else:
            off_df = pd.DataFrame(columns=["Data Inicial", "Data Final", "Tipo"])
    else:
        off_df = st.data_editor(
            pd.DataFrame([{"Tipo": "Férias", "Data Inicial": None, "Data Final": None}]),
            num_rows="dynamic",
            column_config={
                "Data Inicial": st.column_config.DateColumn("Início", format="DD/MM/YYYY"),
                "Data Final": st.column_config.DateColumn("Fim", format="DD/MM/YYYY"),
            },
            use_container_width=True,
            hide_index=True,
        )

    st.divider()

    # ============= BOTÃO GERAR =============
    btn = st.button("🚀 Gerar Cartão de Ponto", type="primary", use_container_width=True)

    if btn:
        try:
            if not base_days_schedule:
                st.error("Defina pelo menos um dia de trabalho!")
                st.stop()

            if input_method == "Colar" and 'text_input' in locals() and text_input.strip():
                off_df = parse_periods_from_text(text_input)

            semanas_final = semanas_sobreaviso if tipo_sobreaviso == TipoSobreaviso.SEMANAS_ESPECIFICAS else []

            with st.spinner("Calculando..."):
                df_card, df_totais, df_irr = build_card(
                    start_date, end_date, base_days_schedule,
                    weekend_mode, weekend_anchor, sat_per_month, sun_per_month,
                    sobreaviso_days,
                    "Fim da jornada → Próx. entrada" if sobreaviso_mode == "Fim da jornada" else "Fixo",
                    so_fixed_start, so_fixed_end,
                    rules_df, off_df, feriados_trabalhados, st.session_state.config,
                    tipo_sobreaviso, semanas_final, sobreaviso_alternado_anchor,
                )

            st.success("✅ Cartão de ponto gerado com sucesso!")

            # Métricas
            col1, col2, col3, col4, col5, col6, col7, col8 = st.columns(8)
            with col1: 
                st.metric("Total Horas", f"{df_card['Total Horas (h)'].sum():.1f}h")
            with col2: 
                st.metric("Horas Líquidas", f"{df_card['Total Horas Líquidas (h)'].sum():.1f}h")
            with col3: 
                st.metric("Horas Extras", f"{df_card['Horas Extras (h)'].sum():.1f}h")
            with col4: 
                st.metric("Horas Diurnas", f"{df_card['Horas Diurnas (h)'].sum():.1f}h")
            with col5: 
                st.metric("Horas Noturnas", f"{df_card['Horas Noturnas 22-5h (h)'].sum():.1f}h")
            with col6: 
                st.metric("Súmula 60", f"{df_card['Horas Súmula 60 (h)'].sum():.1f}h")
            with col7: 
                st.metric("Dom/Fer", f"{df_card['Horas Dom/Fer (h)'].sum():.1f}h")
            with col8: 
                st.metric("Sobreaviso", f"{df_card['Sobreaviso Líquido (h)'].sum():.1f}h")

            # Abas
            tab1, tab2, tab3 = st.tabs(["📋 Cartão de Ponto", "📊 Totais por Mês", "⚠️ Irregularidades"])

            with tab1:
                st.dataframe(df_card, use_container_width=True, hide_index=True)

            with tab2:
                st.dataframe(df_totais, use_container_width=True, hide_index=True)

                # Gráfico
                grafico = {
                    "Diurnas": df_card["Horas Diurnas (h)"].sum(),
                    "Noturnas": df_card["Horas Noturnas 22-5h (h)"].sum(),
                    "Súmula 60": df_card["Horas Súmula 60 (h)"].sum(),
                    "Extras": df_card["Horas Extras (h)"].sum(),
                    "Sobreaviso": df_card["Sobreaviso Líquido (h)"].sum(),
                    "Dom/Fer": df_card["Horas Dom/Fer (h)"].sum(),
                }
                grafico = {k: v for k, v in grafico.items() if v > 0}
                if grafico:
                    st.bar_chart(pd.DataFrame({
                        "Tipo": list(grafico.keys()), 
                        "Horas": list(grafico.values())
                    }), x="Tipo", y="Horas")

            with tab3:
                if not df_irr.empty:
                    for _, row in df_irr.iterrows():
                        st.warning(f"**{row['data'].strftime('%d/%m/%Y')}** - {row['tipo']}: {row['detalhe']}")
                else:
                    st.success("✅ Nenhuma irregularidade encontrada!")

            # Download Excel
            excel_bytes = dataframe_to_excel_bytes(
                df_card, df_totais, df_irr, st.session_state.config,
                tipo_sobreaviso, semanas_final, sobreaviso_alternado_anchor
            )

            st.download_button(
                label="📥 Download Excel Completo",
                data=excel_bytes,
                file_name=f"cartao_ponto_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        except Exception as e:
            st.error(f"Erro ao gerar cartão de ponto: {str(e)}")
            import traceback
            with st.expander("Detalhes do erro"):
                st.code(traceback.format_exc())

